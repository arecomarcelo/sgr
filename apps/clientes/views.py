import locale
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, Optional

import openpyxl
import pandas as pd
import streamlit as st
from st_aggrid import AgGrid, GridOptionsBuilder

from service import DataService


class ClientesReport:
    """
    Classe responsável pelo relatório de Clientes
    Implementa padrão de projeto Facade para interface com Streamlit
    """

    def __init__(self) -> None:
        """Inicializa a classe com configurações básicas"""
        self.configure_locale()
        self.configure_page()
        self.data_service = DataService()

    def configure_locale(self) -> None:
        """Configura localização para formato brasileiro"""
        try:
            locale.setlocale(locale.LC_ALL, "pt_BR.UTF-8")
        except locale.Error:
            locale.setlocale(locale.LC_ALL, "C")

    def configure_page(self) -> None:
        """Configura aparência da página Streamlit"""
        hide_st_style = """
        <style>
        #MainMenu {visibility: hidden;} 
        footer {visibility: hidden;} 
        header {visibility: hidden;} 
        .stDeployButton {visibility: hidden;}
        [data-testid="stStatusWidget"] {visibility: hidden;}
        </style>
        """
        st.markdown(hide_st_style, unsafe_allow_html=True)

    @staticmethod
    @st.cache_data(ttl=300)
    def load_data() -> pd.DataFrame:
        """
        Carrega dados dos Clientes com cache
        Returns:
            pd.DataFrame: DataFrame com os dados dos clientes
        """
        try:
            with st.spinner("Carregando dados de clientes..."):
                return DataService().get_clientes()
        except Exception as e:
            st.error(f"Erro ao carregar dados: {str(e)}")
            return pd.DataFrame()

    def create_grid_options(self, df: pd.DataFrame) -> Dict[str, Any]:
        """
        Configura opções da grade de dados AG-Grid
        Args:
            df (pd.DataFrame): DataFrame com os dados
        Returns:
            dict: Opções de configuração do AG-Grid
        """
        gb_clientes = GridOptionsBuilder.from_dataframe(df)

        gb_clientes.configure_grid_options(
            domLayout="normal",
            enableRangeSelection=True,
            enableCellTextSelection=True,
            suppressRowClickSelection=True,
            enableExcelExport=True,
            enableCsvExport=True,
            onFirstDataRendered="onFirstDataRendered",
            onFilterChanged="onFilterChanged",
        )

        gb_clientes.configure_default_column(
            filter=True,
            cellStyle={"border": "1px solid black"},
            floatingFilter=True,
            resizable=True,
            sortable=True,
        )

        colunas_config = {
            "TipoPessoa": {"headerName": "Tipo", "width": 80},
            "RazaoSocial": {"headerName": "Razão Social", "width": 280},
            "Nome": {"headerName": "Nome", "width": 280},
            "CNPJ": {"headerName": "CNPJ", "width": 200},
            "CPF": {"headerName": "CPF", "width": 120},
            # "Telefone": {"headerName": "Telefone", "width": 120},
            # "Celular": {"headerName": "Celular", "width": 120},
            "Email": {"headerName": "E-mail", "width": 200},
        }

        for campo, config in colunas_config.items():
            if campo in df.columns:
                gb_clientes.configure_column(campo, **config)

        return gb_clientes.build()

    def format_cnpj_cpf(self, value: str) -> str:
        """Formata CNPJ ou CPF para exibição"""
        if pd.isna(value) or value in ["-", ""]:
            return "-"

        value = str(value).strip().replace(".", "").replace("-", "").replace("/", "")

        if len(value) == 11:  # CPF
            return f"{value[:3]}.{value[3:6]}.{value[6:9]}-{value[9:]}"
        elif len(value) == 14:  # CNPJ
            return f"{value[:2]}.{value[2:5]}.{value[5:8]}/{value[8:12]}-{value[12:]}"
        return value

    # def format_phone(self, value: str) -> str:
    #     """Formata telefone para exibição"""
    #     if pd.isna(value) or value in ['-', '']:
    #         return '-'

    #     value = str(value).strip().replace('(', '').replace(')', '').replace('-', '').replace(' ', '')

    #     if len(value) == 10:  # Telefone fixo
    #         return f'({value[:2]}) {value[2:6]}-{value[6:]}'
    #     elif len(value) == 11:  # Celular
    #         return f'({value[:2]}) {value[2:7]}-{value[7:]}'
    #     return value

    def generate_excel(self, df: pd.DataFrame) -> Optional[bytes]:
        """
        Gera arquivo Excel formatado
        Args:
            df (pd.DataFrame): DataFrame com os dados
        Returns:
            bytes: Dados do arquivo Excel ou None em caso de erro
        """
        if df.empty:
            return None

        try:
            output = BytesIO()
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Clientes"

            # Formatar dados antes de exportar
            df_formatted = df.copy()

            # Aplicar formatação para os dados que serão exportados
            if "CNPJ" in df_formatted.columns:
                df_formatted["CNPJ"] = df_formatted["CNPJ"].apply(self.format_cnpj_cpf)
            if "CPF" in df_formatted.columns:
                df_formatted["CPF"] = df_formatted["CPF"].apply(self.format_cnpj_cpf)
            # if 'Telefone' in df_formatted.columns:
            #     df_formatted['Telefone'] = df_formatted['Telefone'].apply(self.format_phone)
            # if 'Celular' in df_formatted.columns:
            #     df_formatted['Celular'] = df_formatted['Celular'].apply(self.format_phone)

            # Escrever cabeçalhos
            for col_idx, column in enumerate(df_formatted.columns, start=1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = column
                cell.font = openpyxl.styles.Font(bold=True)

            # Escrever dados
            for row_idx, row in enumerate(df_formatted.values, start=2):
                for col_idx, value in enumerate(row, start=1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = str(value) if pd.notnull(value) else ""

            # Ajustar largura das colunas
            for col_idx, column in enumerate(df_formatted.columns, start=1):
                max_length = max(
                    df_formatted[column].astype(str).apply(len).max(), len(str(column))
                )
                worksheet.column_dimensions[
                    openpyxl.utils.get_column_letter(col_idx)
                ].width = min(max_length + 2, 50)

            workbook.save(output)
            return output.getvalue()

        except Exception as e:
            st.error(f"Erro ao gerar arquivo Excel: {str(e)}")
            return None

    def run(self, key: Optional[str] = None) -> None:
        """Método principal que executa o relatório"""
        st.title("Relatório de Clientes")

        try:
            # Carrega os dados
            df = self.load_data()

            if df.empty:
                st.warning("Não foram encontrados dados de clientes.")
                return

            # Aplica formatação para exibição
            df_display = df.copy()
            if "CNPJ" in df_display.columns:
                df_display["CNPJ"] = df_display["CNPJ"].apply(self.format_cnpj_cpf)
            if "CPF" in df_display.columns:
                df_display["CPF"] = df_display["CPF"].apply(self.format_cnpj_cpf)
            # if 'Telefone' in df_display.columns:
            #     df_display['Telefone'] = df_display['Telefone'].apply(self.format_phone)
            # if 'Celular' in df_display.columns:
            #     df_display['Celular'] = df_display['Celular'].apply(self.format_phone)

            # Configura as opções do grid
            grid_options = self.create_grid_options(df_display)

            # Renderiza o grid
            with st.spinner("Carregando dados..."):
                grid_response = AgGrid(
                    df_display,
                    gridOptions=grid_options,
                    height=600,
                    fit_columns_on_grid_load=False,
                    theme="alpine",
                    allow_unsafe_jscode=True,
                    reload_data=True,
                    key="clientes_grid",
                )

            # Obtém os dados filtrados da grid
            filtered_data = pd.DataFrame(grid_response["data"])

            # Botão de download do Excel - versão simplificada com download automático
            if not filtered_data.empty:
                excel_bytes = self.generate_excel(filtered_data)
                if excel_bytes:
                    st.download_button(
                        label="📥 Exportar para Excel",
                        data=excel_bytes,
                        file_name=f"relatorio_clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="export_excel",
                    )
            else:
                st.warning("Não há dados filtrados para exportar.")

        except Exception as e:
            st.error(f"Ocorreu um erro ao gerar o relatório: {str(e)}")
            st.exception(e)


def main(key=None):
    report = ClientesReport()
    report.run(key=key)


if __name__ == "__main__":
    main()
