import streamlit as st
from service import DataService
from st_aggrid import AgGrid, GridOptionsBuilder
import pandas as pd
from io import BytesIO
import locale
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from dateutil import parser
import openpyxl


class BoletosReport:
    """
    Classe responsável pelo relatório de Extratos
    Implementa padrão de projeto Facade para interface com Streamlit
    """
    def __init__(self):
        """Inicializa a classe com configurações básicas"""
        self.configure_locale()
        self.configure_page()
        self.data_service = DataService()
        
    def configure_locale(self):
        """Configura localização para formato brasileiro"""
        try:
            locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        except locale.Error:
            locale.setlocale(locale.LC_ALL, 'C')

    def configure_page(self):
        """Configura aparência da página Streamlit"""
        hide_st_style = """
        <style>
        #MainMenu {visibility: hidden;} 
        footer {visibility: hidden;} 
        header {visibility: hidden;} 
        .stDeployButton {visibility: hidden;}
        [data-testid="stStatusWidget"] {visibility: hidden;}
        </style>
        """
        st.markdown(hide_st_style, unsafe_allow_html=True)

    @staticmethod
    @st.cache_data(ttl=300)
    def load_data(data_inicial=None, data_final=None):
        """
        Carrega dados dos Boletos com cache
        Args:
            data_inicial (str): Data inicial no formato YYYY-MM-DD
            data_final (str): Data final no formato YYYY-MM-DD
        Returns:
            pd.DataFrame: DataFrame com os dados dos boletos
        """
        try:
            with st.spinner('Carregando dados...'):
                if not data_inicial:
                    data_inicial = (date.today() - relativedelta(days=1)).strftime('%Y-%m-%d')
                if not data_final:
                    data_final = (date.today() - relativedelta(days=1)).strftime('%Y-%m-%d')
                    
                return DataService().get_boletos_filtrados(
                    data_inicial=data_inicial,
                    data_final=data_final
                )
        except Exception as e:
            st.error(f"Erro ao carregar dados: {str(e)}")
            return pd.DataFrame()

    def format_date(self, value):
        """
        Formata datas para o padrão brasileiro
        Args:
            value: Valor a ser formatado (str, datetime ou date)
        Returns:
            str: Data formatada ou string vazia em caso de erro
        """
        if pd.isna(value):
            return ''
        try:
            if isinstance(value, str):
                try:
                    dt = parser.parse(value)
                    return dt.strftime('%d/%m/%Y %H:%M:%S')
                except ValueError:
                    return datetime.strptime(value, '%Y-%m-%d').strftime('%d/%m/%Y')
            elif isinstance(value, (datetime, date)):
                if isinstance(value, datetime):
                    return value.strftime('%d/%m/%Y %H:%M:%S')
                return value.strftime('%d/%m/%Y')
            return ''
        except (ValueError, TypeError, AttributeError) as e:
            st.error(f"Erro ao formatar data: {str(e)}")
            return str(value)

    def calculate_totals(self, data):
        """
        Calcula totalizadores dos Boletos
        Args:
            data (pd.DataFrame): DataFrame com os dados dos boletos
        Returns:
            dict: Dicionário com os totais calculados
        """
        try:
            if data.empty or 'Status' not in data.columns:
                return {
                    'total_enviados': 0,
                    'total_errados': 0,
                    'total_geral': 0
                }
            
            status_counts = data['Status'].value_counts()
            total_enviados = int(status_counts.get('Enviado', 0))
            total_errados = int(status_counts.get('Errado', 0))
            total_geral = total_enviados + total_errados
            
            return {
                'total_enviados': total_enviados,
                'total_errados': total_errados,
                'total_geral': total_geral
            }
        except Exception as e:
            st.error(f"Erro ao calcular totais: {str(e)}")
            return {
                'total_enviados': 0,
                'total_errados': 0,
                'total_geral': 0
            }

    def create_grid_options(self, df):
        """
        Configura opções da grade de dados AG-Grid
        Args:
            df (pd.DataFrame): DataFrame com os dados
        Returns:
            dict: Opções de configuração do AG-Grid
        """
        gb_boletos = GridOptionsBuilder.from_dataframe(df)
        
        gb_boletos.configure_grid_options(
            domLayout='normal',
            enableRangeSelection=True,
            enableCellTextSelection=True,
            suppressRowClickSelection=True,
            enableExcelExport=True,
            enableCsvExport=True,
            onFirstDataRendered='onFirstDataRendered',
            onFilterChanged='onFilterChanged'
        )

        gb_boletos.configure_default_column(
            filter=True,
            cellStyle={'border': '1px solid black'},
            floatingFilter=True,
            resizable=True,
            sortable=True
        )
        
        colunas_config = {
            "Nome": {"headerName": "Nome"},
            "Boleto": {"headerName": "Boleto"},
            "Vencimento": {
                "headerName": "Vencimento",
                "type": ["dateColumn", "dateColumnFilter"],
                "valueFormatter": "new Date(x).toLocaleDateString('pt-BR')",
                "width": 110
            },
            "envio": {
                "headerName": "Envio",
                # "valueFormatter": "function(params) { return new Date(params.value).toLocaleString('pt-BR', { timeZone: 'UTC' }); }",
                "valueFormatter": "new Date(x).toLocaleDateString('pt-BR')",
                "width": 160
            },
            "Status": {"headerName": "Status"},
        }
        
        for campo, config in colunas_config.items():
            if campo in df.columns:
                gb_boletos.configure_column(campo, **config)
        
        return gb_boletos.build()

    def render_filters(self):
        """
        Renderiza filtros do relatório
        Returns:
            dict: Dicionário com as datas selecionadas
        """
        with st.expander("Filtros", expanded=True):
            col1, col2 = st.columns(2)
            
            with col1:
                data_inicial = st.date_input(
                    "Data Inicial",
                    value=date.today() - relativedelta(days=1),
                    format="DD/MM/YYYY"
                )
            
            with col2:
                data_final = st.date_input(
                    "Data Final",
                    value=date.today() - relativedelta(days=1),
                    format="DD/MM/YYYY"
                )
            
            return {
                'data_inicial': data_inicial.strftime('%Y-%m-%d'),
                'data_final': data_final.strftime('%Y-%m-%d')
            }

    def render_totals(self, totals, df):
        """
        Renderiza totalizadores e botão de download
        Args:
            totals (dict): Dicionário com os totais calculados
            df (pd.DataFrame): DataFrame com os dados
        """
        try:
            container = st.container()
            
            with container:
                col1, col2, col3, col_space, col_button = st.columns([2,2,2,3,1])
                
                with col1:
                    st.metric(
                        label="Total Enviados",
                        value=totals['total_enviados']
                    )
                with col2:
                    st.metric(
                        label="Total Errados",
                        value=totals['total_errados']
                    )
                with col3:
                    st.metric(
                        label="Total Geral",
                        value=totals['total_geral']
                    )
                with col_button:
                    st.write("")
                    if not df.empty:
                        excel_data = self.generate_excel(df)
                        if excel_data:
                            if st.download_button(
                                label="📥 Baixar Excel",
                                data=excel_data,
                                file_name=f"boletos_{date.today().strftime('%d%m%Y')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            ):
                                st.success("Download iniciado!")
                    else:
                        st.warning("Não há dados para exportar.")
        except Exception as e:
            st.error(f"Erro ao renderizar totais: {str(e)}")

    def generate_excel(self, df):
        """
        Gera arquivo Excel formatado
        Args:
            df (pd.DataFrame): DataFrame com os dados
        Returns:
            bytes: Dados do arquivo Excel ou None em caso de erro
        """
        if df.empty:
            return None

        try:
            output = BytesIO()
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Boletos'
            
            df_formatted = df.copy()
            
            for index, row in df_formatted.iterrows():
                if 'envio' in df_formatted.columns:
                    df_formatted.at[index, 'envio'] = self.format_date(row['envio'])
                if 'Vencimento' in df_formatted.columns:
                    df_formatted.at[index, 'Vencimento'] = self.format_date(row['Vencimento'])
            
            for col_idx, column in enumerate(df_formatted.columns, start=1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = column
                
            for row_idx, row in enumerate(df_formatted.values, start=2):
                for col_idx, value in enumerate(row, start=1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = str(value) if pd.notnull(value) else ''
            
            for col_idx, column in enumerate(df_formatted.columns, start=1):
                max_length = max(
                    df_formatted[column].astype(str).apply(len).max(),
                    len(str(column))
                ) + 2
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length
            
            workbook.save(output)
            return output.getvalue()
            
        except Exception as e:
            st.error(f"Erro ao gerar arquivo Excel: {str(e)}")
            return None

    def run(self):
        """Método principal que executa o relatório"""
        st.title("Relatório de Boletos Enviados")
        
        try:
            filtros = self.render_filters()
            df = self.load_data(
                data_inicial=filtros['data_inicial'],
                data_final=filtros['data_final']
            )
            
            if df.empty:
                st.warning("Não foram encontrados dados para os filtros selecionados.")
                return
            
            totals_container = st.container()
            grid_options = self.create_grid_options(df)
            grid_response = AgGrid(
                df,
                gridOptions=grid_options,
                height=800,
                fit_columns_on_grid_load=True,
                theme='alpine',
                allow_unsafe_jscode=True,
                reload_data=True,
                key='grid'
            )
            
            totals = self.calculate_totals(grid_response['data'])
            
            with totals_container:
                self.render_totals(totals, grid_response['data'])
                
            st.markdown("---")
                
        except Exception as e:
            st.error(f"Erro ao carregar os dados: {str(e)}")
            st.exception(e)

def main():
    """Função principal da aplicação"""
    report = BoletosReport()
    report.run()

if __name__ == "__main__":
    main()