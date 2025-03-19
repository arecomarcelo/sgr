import streamlit as st
from service import DataService
from st_aggrid import AgGrid, GridOptionsBuilder
import pandas as pd
from io import BytesIO
import locale
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import openpyxl
from style_utils import apply_default_style

class ExtratosReport:
    """
    Classe responsável pelo relatório de extratos bancários
    Implementa padrão de projeto Facade para interface com Streamlit
    """
    def __init__(self):
        self.configure_locale()
        self.configure_page()
        self.data_service = DataService()
        
    def configure_locale(self):
        """Configura localização para formato brasileiro"""
        try:
            locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        except locale.Error:
            # st.warning("A localidade 'pt_BR.UTF-8' não está disponível. Usando a localidade padrão.")
            locale.setlocale(locale.LC_ALL, 'C')

    def configure_page(self):
        """Configura aparência da página Streamlit"""
        hide_st_style = """
        <style>
        #MainMenu {visibility: hidden;} 
        footer {visibility: hidden;} 
        header {visibility: hidden;} 
        .stDeployButton {visibility: hidden;}
        [data-testid="stStatusWidget"] {visibility: hidden;}
        </style>
        """
        st.markdown(hide_st_style, unsafe_allow_html=True)

    @staticmethod
    @st.cache_data(ttl=300)
    def load_data(data_inicial=None, data_final=None, empresas=None, centros_custo=None):
        """
        Carrega dados dos extratos com cache
        Implementa filtros dinâmicos
        """
        with st.spinner('Carregando dados...'):
            if not data_inicial:
                data_inicial = (date.today() - relativedelta(months=1)).strftime('%Y-%m-%d')
            if not data_final:
                data_final = date.today().strftime('%Y-%m-%d')
                
            return DataService().get_extratos_filtrados(
                data_inicial=data_inicial,
                data_final=data_final,
                empresas=empresas,
                centros_custo=centros_custo
            )

    def format_currency(self, value):
        """Formata valores monetários no padrão brasileiro"""
        if pd.isna(value):
            return ''
        try:
            # Converte para float se for string
            if isinstance(value, str):
                value = float(value.replace('R$', '').replace('.', '').replace(',', '.').strip())
            return f'R$ {value:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        except (ValueError, TypeError):
            return ''

    def format_date(self, value):
        """
        Formata datas para o padrão brasileiro
        """
        if pd.isna(value):
            return ''
        try:
            if isinstance(value, str):
                # Tenta converter string para data
                return datetime.strptime(value, '%Y-%m-%d').strftime('%d/%m/%Y')
            elif isinstance(value, (datetime, date)):
                return value.strftime('%d/%m/%Y')
            return ''
        except (ValueError, TypeError, AttributeError):
            return str(value)

    def calculate_totals(self, data):
        """Calcula totalizadores dos extratos"""
        if data.empty or 'valor' not in data.columns or 'D/C' not in data.columns:
            return {
                'total_credito': 0,
                'total_debito': 0,
                'total_geral': 0
            }

        # Converte valores para numérico, tratando strings formatadas
        def convert_to_numeric(x):
            if isinstance(x, str):
                try:
                    return float(x.replace('R$', '').replace('.', '').replace(',', '.').strip())
                except ValueError:
                    return 0
            return x

        valor = data['valor'].apply(convert_to_numeric)
        debitos = valor[data['D/C'] == 'D'].sum()
        creditos = valor[data['D/C'] == 'C'].sum()
        
        total_geral = creditos - debitos
        
        return {
            'total_credito': float(creditos) if not pd.isna(creditos) else 0,
            'total_debito': float(debitos) if not pd.isna(debitos) else 0,
            'total_geral': float(total_geral) if not pd.isna(total_geral) else 0
        }

    def create_grid_options(self, df):
        """Configura opções da grade de dados AG-Grid"""
        gb = GridOptionsBuilder.from_dataframe(df)
        
        gb.configure_grid_options(
            domLayout='normal',
            enableRangeSelection=True,
            enableCellTextSelection=True,
            suppressRowClickSelection=True,
            enableExcelExport=True,
            enableCsvExport=True,
            onFirstDataRendered='onFirstDataRendered',
            onFilterChanged='onFilterChanged'
        )

        gb.configure_default_column(
            filter=True,
            cellStyle={'border': '1px solid black'},
            floatingFilter=True,
            resizable=True,
            sortable=True
        )
        
        colunas_config = {
            "banco": {"headerName": "Banco", "width": 100},
            "agencia": {"headerName": "Agência", "width": 100},
            "conta_corrente": {"headerName": "Conta Corrente", "width": 130},
            "data": {
                "headerName": "Data",
                "type": ["dateColumn", "dateColumnFilter"],
                "valueFormatter": "new Date(x).toLocaleDateString('pt-BR')",
                "width": 110
            },
            "documento": {"headerName": "Documento", "width": 120},
            "descricao": {"headerName": "Descrição", "width": 300},
            "valor": {
                "headerName": "Valor",
                "type": ["numericColumn", "numberColumnFilter"],
                "valueFormatter": "'R$ ' + x.toLocaleString('pt-BR', {minimumFractionDigits: 2, maximumFractionDigits: 2})",
                "width": 130
            },
            "D/C": {"headerName": "D/C", "width": 80},
            "empresa": {"headerName": "Empresa", "width": 200},
            "centrocusto": {"headerName": "Centro de Custo", "width": 200}
        }
        
        for campo, config in colunas_config.items():
            if campo in df.columns:
                gb.configure_column(campo, **config)
        
        return gb.build()

    def render_filters(self):
        """Renderiza filtros do relatório"""
        with st.expander("Filtros", expanded=True):
            col1, col2 = st.columns(2)
            
            with col1:
                data_inicial = st.date_input(
                    "Data Inicial",
                    value=date.today() - relativedelta(months=12),
                    format="DD/MM/YYYY"
                )
            
            with col2:
                data_final = st.date_input(
                    "Data Final",
                    value=date.today(),
                    format="DD/MM/YYYY"
                )
            
            return {
                'data_inicial': data_inicial.strftime('%Y-%m-%d'),
                'data_final': data_final.strftime('%Y-%m-%d')
            }

    def render_totals(self, totals, df):
        """Renderiza totalizadores e botão de download"""
        container = st.container()
        
        with container:
            col1, col2, col3, col_space, col_button = st.columns([2,2,2,3,1])
            
            with col1:
                st.metric(
                    label="Total Créditos",
                    value=self.format_currency(totals['total_credito'])
                )
            with col2:
                st.metric(
                    label="Total Débitos",
                    value=self.format_currency(totals['total_debito'])
                )
            with col3:
                st.metric(
                    label="Total Geral",
                    value=self.format_currency(totals['total_geral'])
                )
            with col_button:
                st.write("")
                if not df.empty:
                    excel_data = self.generate_excel(df)
                    if excel_data:
                        if st.download_button(
                            label="📥 Baixar Excel",
                            data=excel_data,
                            file_name=f"extratos_{date.today().strftime('%d%m%Y')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        ):
                            st.success("Download iniciado!")
                else:
                    st.warning("Não há dados para exportar.")

    def generate_excel(self, df):
        """Gera arquivo Excel formatado"""
        if df.empty:
            return None

        try:
            output = BytesIO()
            
            # Criar um novo workbook e selecionar a planilha ativa
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Extratos'
            
            # Preparar dados formatados
            df_formatted = df.copy()
            
            # Formatação dos valores e datas
            for index, row in df_formatted.iterrows():
                # Formata valor
                if 'valor' in df_formatted.columns:
                    df_formatted.at[index, 'valor'] = self.format_currency(row['valor'])
                
                # Formata data
                if 'data' in df_formatted.columns:
                    df_formatted.at[index, 'data'] = self.format_date(row['data'])
            
            # Escrever cabeçalhos
            for col_idx, column in enumerate(df_formatted.columns, start=1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = column
                
            # Escrever dados
            for row_idx, row in enumerate(df_formatted.values, start=2):
                for col_idx, value in enumerate(row, start=1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = str(value) if pd.notnull(value) else ''
            
            # Ajustar largura das colunas
            for col_idx, column in enumerate(df_formatted.columns, start=1):
                max_length = max(
                    df_formatted[column].astype(str).apply(len).max(),
                    len(str(column))
                ) + 2
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length
            
            # Salvar o workbook
            workbook.save(output)
            return output.getvalue()
            
        except Exception as e:
            st.error(f"Erro ao gerar arquivo Excel: {str(e)}")
            return None

    def run(self, key=None):
        """Método principal que executa o relatório"""
        st.title("Relatório de Extratos Bancários")
        
        try:
            # Renderizar filtros
            filtros = self.render_filters()
            
            # Carregar dados com filtros
            df = self.load_data(
                data_inicial=filtros['data_inicial'],
                data_final=filtros['data_final']
            )
            
            if df.empty:
                st.warning("Não foram encontrados dados para os filtros selecionados.")
                return
            
            # Container para totalizadores
            totals_container = st.container()
            
            # Usar st.empty() como placeholder para o grid
            grid_placeholder = st.empty()
            
            with st.spinner('Carregando grid...'):
                # Renderizar o grid dentro do placeholder
                with grid_placeholder:
                    grid_options = self.create_grid_options(df)
                    grid_response = AgGrid(
                        df,
                        gridOptions=grid_options,
                        height=800,
                        fit_columns_on_grid_load=True,
                        theme='alpine',
                        allow_unsafe_jscode=True,
                        reload_data=True,
                        key=f'grid_{key}'  # Adicionando a chave única
                    )
            
            # Calcular e exibir totalizadores
            totals = self.calculate_totals(grid_response['data'])
            
            with totals_container:
                self.render_totals(totals, grid_response['data'])
                
            st.markdown("---")
                
        except Exception as e:
            st.error(f"Erro ao carregar os dados: {str(e)}")
            st.exception(e)

def main(key=None):
    report = ExtratosReport()
    report.run(key=key)

if __name__ == "__main__":
    main()